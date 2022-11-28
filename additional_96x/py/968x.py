# same as 956 just change the them field


# there is a second file with the following columns.
# filename
# Month_Yr
# all_pages
# date
# policy_prop
# policy_prop_Submitter
# policy_prop_future_date
# others
# theme
# count
# Location
# policy_prop_type
# all_pages2



import sys, os, json, gzip
cwd = os.getcwd().replace('\\', '/') # the cwd is the location where python starts, in this case, it is C:\Personal\Virtual_Server\PHPWeb\ml_text\python
# add the location of the py common module file 
# location_commonmodules = cwd + '/pyflaskapps/common_pymodules'
# sys.path.insert(1, location_commonmodules) # add the location at the top of the sys.path, thus, by default, the py script file spacy_modules will be loaded from there
# from modules_common import * 
# from modules_spacy import *


tmpdir= '/additional_96x/data/history/'
tmpoutdir  = '/additional_96x/data/history/out/' 


# 1. load the selected themes
# selected_themes_file = cwd + "/pyflaskapps/data/output/new_policyanalysis/07_themes/selected_themes.json"
# openedfile0 = open(selected_themes_file, encoding="UTF-8")
# selected_themes_dict = json.load(openedfile0) 
# selected_themes_ls = selected_themes_dict['data']['selected_themes']
selected_themes_ls = ["harm_reduction",            "housing",            "mental_health",            "youth_children"]
selected_themes_ls = selected_themes_ls + ['others']
print(32, selected_themes_ls)

# 2. load the src file (note: it is not a textacy obj, just a plain json)
srcfilepath = cwd + tmpoutdir + '/allsents_same_structure_as_with_matched_phrases_of_selected_sections.json.gz'
# srcfilepath = cwd + '/pyflaskapps/data/output/new_policyanalysis/04a_corpus/sents_with_matched_phrases.json.json.gz'
with gzip.open(srcfilepath, 'r') as fin:        # 4. gzip
    json_bytes = fin.read()                     # 3. bytes (i.e. UTF-8)
# convert bytes to string
json_str = json_bytes.decode('utf-8')  
# convert string to json          # 2. string (i.e. JSON)
jsonobj = json.loads(json_str) # like {sent_text:..., meta:{...}, matched_phrases_themes:[phrase:..., themes:p[]]}
sents_ls = jsonobj['data']

def make_distinct_recs(sents_ls):
    test_ls =[]
    distinct_ls =[]
    for x in sents_ls:
        # keystr = x['sent_text']+'|'+x['meta']['city']+ '|' +x['meta']['date']+ '|' +x['meta']['file']+ '|' + str(x['meta']['sectionid'])+ '|' + str(x['meta']['subsectionid'])
        keystr = json.dumps(x)
        if (keystr not in test_ls):
            test_ls.append(keystr)
            distinct_ls.append(x)
    return distinct_ls
print(55, 'length before remove dups', len(sents_ls))
sents_ls = make_distinct_recs(sents_ls)
print(57, 'length after remove dups', len(sents_ls))
#########################################################################


# read the subsections into a list
# subsections_filepath = r'C:\Users\syao2\AppData\Local\MyWorks\js\policy_analysis\data\out\council_meetings\corpus\meetingsubsections.json.gz'.replace('\\', '/')
subsections_filepath = cwd + tmpdir + '/meetingsubsections.json.gz'

# print(corpusfilepath)
with gzip.open(subsections_filepath, 'r') as fin:        
    json_bytes = fin.read()      
# convert bytes to string
json_str = json_bytes.decode('utf-8')  
# convert string to json  
subsections_ls = json.loads(json_str)   #like {text: '...', meta: {subsection: ..., }}

# need to transform the subsections_ls into a dictionary so that the subsection text can be looked up by the key of file-city-date-section-subsectionid
subsections_dict= {}
for x in subsections_ls:
    keystr = x['meta']['city']+ '|' +x['meta']['date']+ '|' +x['meta']['file']+ '|' +str(x['meta']['sectionid'])+ '|' + str(x['meta']['subsectionid'])
    try:
        subsections_dict[keystr]
    except:
        subsections_dict[keystr] = x['text']

# get the subsection text in a dictionary
def get_policy_prop(srcdict, filename, city, date, sectionid, subsectionid):
    keystr = city + '|' + date+ '|' + filename +  '|' + str(sectionid) + '|' + str(subsectionid)
    return srcdict[keystr]

#########################################################################


# read the original jsons into a list


# subsections_filepath = r'C:\Users\syao2\AppData\Local\MyWorks\js\policy_analysis\data\out\council_meetings\corpus\meetingsubsections.json.gz'.replace('\\', '/')
subsections_filepath = cwd + tmpdir + '/meetingsubsections.json.gz'
# print(corpusfilepath)
with gzip.open(subsections_filepath, 'r') as fin:        
    json_bytes = fin.read()      
# convert bytes to string
json_str = json_bytes.decode('utf-8')  
# convert string to json  
subsections_ls = json.loads(json_str)   #like {text: '...', meta: {subsection: ..., }}

# need to transform the subsections_ls into a dictionary of minutes
# 1. text of each section should be concatenated so as to have a text body for the whole file
# 2. the whole text body of a file should be made into the minutes_dict with the city date and name as the key
minutes_dict= {}
keystrs_ls=[]
for x in subsections_ls:
    keystr = x['meta']['city']+ '|' + x['meta']['date']+ '|' +x['meta']['file']
    if keystr not in keystrs_ls:
        keystrs_ls.append(keystr)
        minutes_dict[keystr] =""
    minutes_dict[keystr] += '\n====================' + x['meta']['section'] + '__' +  str(x['meta']['subsectionid']) + '\n' + x['text']


# get the subsection text in a dictionary
def get_all_pages(srcdict, filename, city, date):
    keystr = city + '|' + date+ '|' + filename 
    return srcdict[keystr]

#########################################################################

# print(minutes_dict[keystrs_ls[-1]])

new_subsections_ls =[]
uniquestrs_ls =[]

ukey_themes_dict ={} # like ukey: [theme1, 2, ...]
for x in sents_ls:
    meta_dict = x['meta']
    subsectionid = x['meta']['subsectionid']
    future_action_dates_str= meta_dict['future_action_dates']
    if future_action_dates_str =="":
        policy_prop_future_date_ls=[]
    else:
        policy_prop_future_date_ls = future_action_dates_str.split('|') # can be like 'future!2022-0x-xx'

    policy_prop_future_date =""
    policy_prop_future_action="no"

    if future_action_dates_str == 'future':
        policy_prop_future_action = "yes"
    elif len(policy_prop_future_date_ls) > 0:
        policy_prop_future_action = "yes"
        jj =0               
        for x3 in policy_prop_future_date_ls:
            if x3 != 'future': # skip if the value is 'future'
                jj +=1
                policy_prop_future_date = x3
                break #stop at the first date
    for y in x['new_phrase_themes']:
        phrase = y['phrase']
        for theme in y['themes']:
            if theme not in selected_themes_ls: #skip themes that are not in the selected four
                continue
            #  filename => name of the meeting minute
            filename = meta_dict['file']
            #  Month_Yr => a combination of month and year of the meeting date
            Month_Yr = meta_dict['date'][5:7] + '_' + meta_dict['date'][0:4]
            
            #  date => the meeting date
            date = meta_dict['date']

            # all_pages
            # need to make it from the json file by subsection and to make it by order 
            city = meta_dict['city']
            all_pages = get_all_pages(minutes_dict, filename, city, date)

            # section
            sectionid = meta_dict['sectionid']
            sectiontitle = meta_dict['section']
            
            # policy_prop
            # need to have it from the file by subsection (not by sentence)
            policy_prop = get_policy_prop(subsections_dict, filename, city, date, sectionid, subsectionid)

            #  policy_prop_Submitter => the person who moved the proposal
            policy_prop_Submitter = meta_dict['moved']

             #  others => leave blank
            others=""

            # theme
            # theme = theme
            #  count => always 1	
            count = 1

            #  Location => city 
            Location = meta_dict['city']

            #  policy_prop_type => section title
            policy_prop_type = meta_dict['section']

            # all_pages2
            all_pages2 =""

            # it is to find a theme-matched subsection, not for each sentences
            # therefore, once a them is identified, the subsection should be remembered as linking to that theme, 
            # like file - section - subsectionid - theme as a unique record for each subsection
            uniquestr = filename + '|' + sectiontitle + '|' + str(subsectionid) + '|' + theme
            if uniquestr in uniquestrs_ls:
                continue
            else:
                uniquestrs_ls.append(uniquestr)

            policy_prop_seconded = meta_dict['seconded']

            #  others => leave blank
            others=""
            #  theme => theme of the theme-related phrase that appears in the sentence
            # theme = theme
            #  count => always 1	
            count = 1
            #  Location => city 
            Location = meta_dict['city']
            #  policy_prop_type => section title
            policy_prop_type = meta_dict['section']
            #  count_overall => number of time a theme-related phrase appear in all meeting minutes
            count_overall = ""
            try:
                if (len(meta_dict['opposedresult']) > 0):
                    vote_result = meta_dict['carryresult'] + '; Opposed: ' + meta_dict['opposedresult']
                else:
                    vote_result = meta_dict['carryresult']
            except:
                vote_result = meta_dict['carryresult']
            # apeend to the new_subsections_ls
            ukey = date + '|' + Location + '|' + filename + '|' + str(sectionid) + '|' + str(subsectionid) 
            try:
                if theme not in ukey_themes_dict[ukey]:
                    ukey_themes_dict[ukey].append(theme)
            except:
                ukey_themes_dict[ukey] = [theme]
            new_subsections_ls.append({
                'filename':filename,
                'Month_Yr':Month_Yr,
                'all_pages':all_pages,
                'date':date,
                'policy_prop':policy_prop,
                'policy_prop_Submitter':policy_prop_Submitter,
                'policy_prop_future_date':policy_prop_future_date,
                'others':others,
                'theme':"",
                'count':count,
                'Location':Location,
                'policy_prop_type':policy_prop_type,                
                'all_pages2':all_pages2,
                'sectionid':sectionid,
                'subsectionid':subsectionid,
                'policy_prop_seconded':policy_prop_seconded,
                'policy_prop_future_action':policy_prop_future_action,
                'vote_result':vote_result,
                "ukey": ukey,
                'theme2':theme,
            })

# repeat to make the themecat (if count theme2 > 1)
new_new_subsections_ls =[] # excluding records with theme = "others" while x['subsection_themes'] has > 1 themes
for x in new_subsections_ls:
    x['subsection_themes'] = ukey_themes_dict[x['ukey']]
    if (len(x['subsection_themes'])==1) & (x['subsection_themes'][0] == 'others'):
        x['theme'] = 'unclassified'
    else:
         x['theme'] = x['theme2']
    # exclude records that 'theme2 is others, while  x['subsection_themes'] has more than 1 themes
    if (x['theme2'] == 'others') & (len(x['subsection_themes']) >1):
        pass
    else:
        new_new_subsections_ls.append(x)


# save the new_subsections_ls to json
targetfile = cwd + tmpoutdir + 'themesdata_allcities_968.json' 
with open(targetfile, 'w') as f:
    json.dump(new_new_subsections_ls, f)



# save as a xlsx file
def save_excel(srcjson, targetfile, sheet):
    import pandas  as pd
    from openpyxl import Workbook as open_workbook
    from openpyxl import load_workbook

    keys = list(srcjson[0].keys()) # srcjson is like [{keyA:.., keyB:..}, {...}

    # transform the srcjson [{keyA:.., keyB:..}, {...} to dict1 of {keyA: [], keyB:[]} 
    dict1 ={}
    for key in keys:
        dict1[key] =[]
        for x in srcjson:
            dict1[key].append(x[key]) 

    df = pd.DataFrame(dict1)

    xlexist = os.path.exists(targetfile)
    # # print(1711, xlexist)
    if xlexist == True:
        book = load_workbook(targetfile)
    else:
        book = open_workbook()
        book.save(targetfile)


    sheetname = "Sheet"
    try:
        sheet_to_del = book.get_sheet_by_name(sheetname)
        book.remove_sheet(sheet_to_del)
    except:
        print ("no such sheet called %(v1)s in the excle file %(v2)s)" % {'v1': sheetname, 'v2': targetfile})
        pass
    # book = load_workbook(xlsxfile)
    writer = pd.ExcelWriter(targetfile, engine='openpyxl') 
    writer.book=book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets) # if not loading the existing sheets, ExcelWriter will remove all existing sheets and create new ones
    
    # must delete all contents of the existing cells, otherwise it might not get updated (e.g., the new data has 10 rows, while the existing data in xls sheet has 12 rows, the last two rows won't be updated -- the file now has 10 rows of new data and 2 rows of old data )
    print('=== removing exsiting contents in sheet...', sheet)
    targetws = book[sheet]
    for row in targetws['A2:Z50000']:
        for cell in row:
            cell.value = None
    print('=== done removing ')
    df.to_excel(writer, sheet, index=False)
    writer.save()

####################################
targetfile = cwd + tmpoutdir +  "/968_themesdata_allcities_auto.xlsx" # !!! 23 tableau
save_excel(new_new_subsections_ls, targetfile, "themesdata")