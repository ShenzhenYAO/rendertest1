# same as 955, just use new_phrase_themes instead


#  policy_prop => subsection, 
#  word_cloud_hit_sentence => the sentence containing a theme-related phrase
#  wrdcl => the theme-related phrase that appears in the sentence
#  filename => name of the meeting minute
#  Month_Yr => a combination of month and year of the meeting date
#  date => the meeting date
#  policy_prop_Submitter => the person who moved the proposal
#  policy_prop_future_date	=> the future action date
#  others => leave blank
#  theme => theme of the theme-related phrase that appears in the sentence
#  count => always 1	
#  Location => city 
#  policy_prop_type => section title
#  count_overall => number of time a theme-related phrase appear in all meeting minutes

# based on cwd + '/pyflaskapps/data/output/new_policyanalysis/04a_corpus/sents_with_matched_phrases_of_selected_sections.json.gz'
# like {sent_text: ... meta: {city:..., ...}, matched_phrases_themes: [{phrase:.., themes:[]}, {}]} 

import sys, os, json, gzip
cwd = os.getcwd().replace('\\', '/') # the cwd is the location where python starts, in this case, it is C:\Personal\Virtual_Server\PHPWeb\ml_text\python
# add the location of the py common module file 
# location_commonmodules = cwd + '/pyflaskapps/common_pymodules'
# sys.path.insert(1, location_commonmodules) # add the location at the top of the sys.path, thus, by default, the py script file spacy_modules will be loaded from there
# from modules_common import * 
# from modules_spacy import *


tmpdir= '/additional_96x/data/history/'
tmpoutdir  = '/additional_96x/data/history/out/' 

# 1. load the selected themes
# selected_themes_file = cwd + tmpoutdir+  "/pyflaskapps/data/output/new_policyanalysis/07_themes/selected_themes.json"
# openedfile0 = open(selected_themes_file, encoding="UTF-8")
# selected_themes_dict = json.load(openedfile0) 
# selected_themes_ls = selected_themes_dict['data']['selected_themes']
selected_themes_ls = ["harm_reduction",            "housing",            "mental_health",            "youth_children"]
selected_themes_ls = selected_themes_ls + ['others']
print(32, selected_themes_ls)

# 2. load the sentences file (note: it is not a textacy obj, just a plain json)
srcfilepath = cwd + tmpoutdir + '/allsents_same_structure_as_with_matched_phrases_of_selected_sections.json.gz'
# srcfilepath = cwd + '/pyflaskapps/data/output/new_policyanalysis/04a_corpus/sents_with_matched_phrases.json.json.gz'
with gzip.open(srcfilepath, 'r') as fin:        # 4. gzip
    json_bytes = fin.read()                     # 3. bytes (i.e. UTF-8)
# convert bytes to string
json_str = json_bytes.decode('utf-8')  
# convert string to json          # 2. string (i.e. JSON)
jsonobj = json.loads(json_str) # like {sent_text:..., meta:{...}, matched_phrases_themes:[phrase:..., themes:p[]]}
sents_ls = jsonobj['data']

print(45, 'length of sents_ls from allsents_same_structure_as_with_matched_phrases_of_selected_sections.json', len(sents_ls))


def make_distinct_recs(sents_ls):
    test_ls =[]
    distinct_ls =[]
    for x in sents_ls:
        # keystr = x['sent_text']+'|'+x['meta']['city']+ '|' +x['meta']['date']+ '|' +x['meta']['file']+ '|' + str(x['meta']['sectionid'])+ '|' + str(x['meta']['subsectionid'])
        keystr = json.dumps(x)
        if (keystr not in test_ls):
            test_ls.append(keystr)
            distinct_ls.append(x)
    return distinct_ls

# 3. load the subsections into a list (so as to have the subsections)

# subsections_filepath = r'C:\Users\syao2\AppData\Local\MyWorks\js\policy_analysis\data\out\council_meetings\corpus\meetingsubsections.json.gz'.replace('\\', '/')
subsections_filepath = cwd + tmpdir + '/meetingsubsections.json.gz'

# print(corpusfilepath)
with gzip.open(subsections_filepath, 'r') as fin:        
    json_bytes = fin.read()      
# convert bytes to string
json_str = json_bytes.decode('utf-8')  
# convert string to json  
subsections_ls = json.loads(json_str)   #like {text: '...', meta: {subsection: ..., }}

# need to transform the subsections_ls into a dictionary so that the subsection text can be looked up by the key of file-city-date-section-subsectionid
subsections_dict= {}
for x in subsections_ls:
    keystr = x['meta']['city']+ '|' +x['meta']['date']+ '|' +x['meta']['file']+ '|' + str(x['meta']['sectionid'])+ '|' + str(x['meta']['subsectionid'])
    try:
        subsections_dict[keystr]
    except:
        subsections_dict[keystr] = x['text']

# get the subsection text in a dictionary
def get_policy_prop(srcdict, filename, city, date, sectionid, subsectionid):
    keystr = city + '|' + date+ '|' + filename +  '|' + str(sectionid) + '|' + str(subsectionid)
    return srcdict[keystr]

new_sents_ls =[]
distinct_ls = make_distinct_recs(sents_ls) # not sure why there are dup records in sents_ls! remove the dups
for x in sents_ls:
    meta_dict = x['meta']

    # the following can be done at sentence level:
    #  word_cloud_hit_sentence => the sentence containing a theme-related phrase
    word_cloud_hit_sentence= x['sent_text']
    #  filename => name of the meeting minute
    filename = meta_dict['file']
    #  Month_Yr => a combination of month and year of the meeting date
    Month_Yr = meta_dict['date'][5:7] + '_' + meta_dict['date'][0:4]
    #  date => the meeting date
    date = meta_dict['date']
    #  policy_prop => subsection, 
    city = meta_dict['city']
    sectionid = meta_dict['sectionid']
    section = meta_dict['section']
    subsectionid = meta_dict['subsectionid']
    policy_prop = get_policy_prop(subsections_dict, filename, city, date, sectionid, subsectionid)
    #  policy_prop_Submitter => the person who moved the proposal
    policy_prop_Submitter = meta_dict['moved']
    policy_prop_seconded = meta_dict['seconded']

    #  policy_prop_future_date	=> the future action date
    policy_prop_future_date = meta_dict['future_action_dates'].split('|')[0]
    policy_prop_future_action="no"
    if meta_dict['future_action_dates'] == 'future':
        policy_prop_future_action = "yes"
        policy_prop_future_date = ""
    elif len(meta_dict['future_action_dates']) > 0:
        policy_prop_future_action = "yes"
    #  others => leave blank
    others=""
    #  theme => theme of the theme-related phrase that appears in the sentence
    # theme = theme
    #  count => always 1	
    count = 1
    #  Location => city 
    Location = meta_dict['city']
    #  policy_prop_type => section title
    policy_prop_type = meta_dict['section']
    #  count_overall => number of time a theme-related phrase appear in all meeting minutes
    count_overall = ""
    try:
        if (len(meta_dict['opposedresult']) > 0):
            vote_result = meta_dict['carryresult'] + '; Opposed: ' + meta_dict['opposedresult']
        else:
            vote_result = meta_dict['carryresult']
    except:
        vote_result = meta_dict['carryresult']
    ukey = date + '|' + Location + '|' + filename + '|' + str(sectionid) + '|' + str(subsectionid) 
    for y in x['new_phrase_themes']:
        phrase = y['phrase']
        for theme in y['themes']:
            if theme not in selected_themes_ls:
                continue

            #  wrdcl => the theme-related phrase that appears in the sentence
            wrdcl = phrase
            theme_cat = theme
            if theme == 'others':
                theme_cat = 'unclassifierd'
            # apeend to the new_sents_ls
            new_sents_ls.append({
                'policy_prop':policy_prop,
                'word_cloud_hit_sentence':word_cloud_hit_sentence,
                'wrdcl':wrdcl,
                'filename':filename,
                'Month_Yr':Month_Yr,
                'date':date,
                'policy_prop_Submitter':policy_prop_Submitter,
                'policy_prop_future_date':policy_prop_future_date,
                'others':others,
                'theme':theme_cat,
                'count':count,
                'Location':Location,
                'policy_prop_type':policy_prop_type,
                'count_overall':count_overall,
                'sectionid': sectionid,
                'subsectionid': subsectionid,
                'policy_prop_seconded':policy_prop_seconded,
                'policy_prop_future_action':policy_prop_future_action,
                'vote_result':vote_result,
                'theme2':theme, 
                'ukey': ukey
            })

# save the new_sents_ls to json
print(167, 'number of records in wordcloud.json', len(new_sents_ls))

# remove dups
nodup_new_sents_ls = make_distinct_recs(new_sents_ls)
print(171, 'number of distinct records in wordcloud.json', len(nodup_new_sents_ls))

targetfile = cwd + tmpoutdir + '/wordcloud_967.json' 
with open(targetfile, 'w') as f:
    json.dump(nodup_new_sents_ls, f)


# save as a xlsx file
def save_excel(srcjson, targetfile, sheet):
    import pandas  as pd
    from openpyxl import Workbook as open_workbook
    from openpyxl import load_workbook

    keys = list(srcjson[0].keys()) # srcjson is like [{keyA:.., keyB:..}, {...}

    # transform the srcjson [{keyA:.., keyB:..}, {...} to dict1 of {keyA: [], keyB:[]} 
    dict1 ={}
    for key in keys:
        dict1[key] =[]
        for x in srcjson:
            dict1[key].append(x[key]) 

    df = pd.DataFrame(dict1)

    xlexist = os.path.exists(targetfile)
    # # print(1711, xlexist)
    if xlexist == True:
        book = load_workbook(targetfile)
    else:
        book = open_workbook()
        book.save(targetfile)


    sheetname = "Sheet"
    try:
        sheet_to_del = book.get_sheet_by_name(sheetname)
        book.remove_sheet(sheet_to_del)
    except:
        print ("no such sheet called %(v1)s in the excle file %(v2)s)" % {'v1': sheetname, 'v2': targetfile})
        pass
    # book = load_workbook(xlsxfile)
    writer = pd.ExcelWriter(targetfile, engine='openpyxl') 
    writer.book=book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets) # if not loading the existing sheets, ExcelWriter will remove all existing sheets and create new ones
    
    # must delete all contents of the existing cells, otherwise it might not get updated (e.g., the new data has 10 rows, while the existing data in xls sheet has 12 rows, the last two rows won't be updated -- the file now has 10 rows of new data and 2 rows of old data )
    print('=== removing exsiting contents in sheet...', sheet)
    targetws = book[sheet]
    for row in targetws['A2:Z160000']:
        for cell in row:
            cell.value = None
    print('removing done. ')
    df.to_excel(writer, sheet, index=False)
    writer.save()

####################################
targetfile = cwd + tmpoutdir +  "/967_wordclouddata_allcities_auto.xlsx" # !!!! 23_tableau
save_excel(nodup_new_sents_ls, targetfile, "wordcloud")